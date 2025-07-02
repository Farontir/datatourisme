"""
Service d'export des données touristiques
"""
import csv
import json
import os
from datetime import datetime
from typing import Tuple, List, Dict
from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """Service centralisé pour l'export des données"""
    
    def __init__(self):
        self.export_dir = getattr(settings, 'EXPORT_DIR', os.path.join(settings.MEDIA_ROOT, 'exports'))
        self._ensure_export_dir()
    
    def _ensure_export_dir(self):
        """S'assurer que le répertoire d'export existe"""
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir, exist_ok=True)
    
    def export_to_csv(self, queryset, language='fr') -> Tuple[str, int]:
        """
        Exporte les ressources en format CSV
        
        Args:
            queryset: QuerySet des ressources à exporter
            language: Langue pour les données multilingues
            
        Returns:
            Tuple (file_path, file_size)
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'resources_export_{timestamp}.csv'
        file_path = os.path.join(self.export_dir, filename)
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # En-têtes
                headers = [
                    'ID Ressource',
                    'Nom',
                    'Description',
                    'Types',
                    'Catégories',
                    'Ville',
                    'Code Postal',
                    'Adresse',
                    'Latitude',
                    'Longitude',
                    'Date Création',
                    'Dernière Mise à Jour',
                    'Actif',
                    'Contact Téléphone',
                    'Contact Email',
                    'Contact Site Web'
                ]
                writer.writerow(headers)
                
                # Données
                for resource in queryset:
                    row = [
                        resource.resource_id,
                        self._get_localized_field(resource, 'name', language),
                        self._get_localized_field(resource, 'description', language),
                        ', '.join(resource.resource_types or []),
                        ', '.join(resource.categories or []),
                        resource.city or '',
                        resource.postal_code or '',
                        resource.address or '',
                        resource.location.y if resource.location else '',
                        resource.location.x if resource.location else '',
                        resource.creation_date.strftime('%Y-%m-%d') if resource.creation_date else '',
                        resource.last_update.strftime('%Y-%m-%d') if resource.last_update else '',
                        'Oui' if resource.is_active else 'Non',
                        getattr(resource, 'contact_phone', '') or '',
                        getattr(resource, 'contact_email', '') or '',
                        getattr(resource, 'contact_website', '') or ''
                    ]
                    writer.writerow(row)
            
            file_size = os.path.getsize(file_path)
            logger.info(f"Export CSV créé: {file_path} ({file_size} bytes)")
            
            return file_path, file_size
            
        except Exception as e:
            logger.error(f"Erreur export CSV: {e}")
            raise
    
    def export_to_excel(self, queryset, language='fr') -> Tuple[str, int]:
        """
        Exporte les ressources en format Excel
        
        Args:
            queryset: QuerySet des ressources à exporter
            language: Langue pour les données multilingues
            
        Returns:
            Tuple (file_path, file_size)
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'resources_export_{timestamp}.xlsx'
        file_path = os.path.join(self.export_dir, filename)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ressources Touristiques"
            
            # Style pour les en-têtes
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # En-têtes
            headers = [
                'ID Ressource',
                'Nom',
                'Description',
                'Types',
                'Catégories',
                'Ville',
                'Code Postal',
                'Adresse',
                'Latitude',
                'Longitude',
                'Date Création',
                'Dernière Mise à Jour',
                'Actif',
                'Contact Téléphone',
                'Contact Email',
                'Contact Site Web'
            ]
            
            # Ajouter les en-têtes avec style
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Données
            for row_idx, resource in enumerate(queryset, 2):
                data = [
                    resource.resource_id,
                    self._get_localized_field(resource, 'name', language),
                    self._get_localized_field(resource, 'description', language),
                    ', '.join(resource.resource_types or []),
                    ', '.join(resource.categories or []),
                    resource.city or '',
                    resource.postal_code or '',
                    resource.address or '',
                    resource.location.y if resource.location else '',
                    resource.location.x if resource.location else '',
                    resource.creation_date if resource.creation_date else '',
                    resource.last_update if resource.last_update else '',
                    'Oui' if resource.is_active else 'Non',
                    getattr(resource, 'contact_phone', '') or '',
                    getattr(resource, 'contact_email', '') or '',
                    getattr(resource, 'contact_website', '') or ''
                ]
                
                for col_idx, value in enumerate(data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Ajouter une feuille de métadonnées
            meta_ws = wb.create_sheet("Métadonnées")
            meta_ws.append(['Export généré le', timezone.now().strftime('%d/%m/%Y à %H:%M')])
            meta_ws.append(['Nombre de ressources', queryset.count()])
            meta_ws.append(['Langue', language])
            meta_ws.append(['Filtres appliqués', 'Ressources actives uniquement'])
            
            wb.save(file_path)
            
            file_size = os.path.getsize(file_path)
            logger.info(f"Export Excel créé: {file_path} ({file_size} bytes)")
            
            return file_path, file_size
            
        except Exception as e:
            logger.error(f"Erreur export Excel: {e}")
            raise
    
    def export_to_json(self, queryset, language='fr') -> Tuple[str, int]:
        """
        Exporte les ressources en format JSON
        
        Args:
            queryset: QuerySet des ressources à exporter
            language: Langue pour les données multilingues
            
        Returns:
            Tuple (file_path, file_size)
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'resources_export_{timestamp}.json'
        file_path = os.path.join(self.export_dir, filename)
        
        try:
            export_data = {
                'metadata': {
                    'export_date': timezone.now().isoformat(),
                    'language': language,
                    'total_resources': queryset.count(),
                    'version': '1.0'
                },
                'resources': []
            }
            
            for resource in queryset:
                resource_data = {
                    'resource_id': resource.resource_id,
                    'name': self._get_localized_field(resource, 'name', language),
                    'description': self._get_localized_field(resource, 'description', language),
                    'resource_types': resource.resource_types or [],
                    'categories': resource.categories or [],
                    'location': {
                        'city': resource.city,
                        'postal_code': resource.postal_code,
                        'address': resource.address,
                        'coordinates': {
                            'latitude': resource.location.y if resource.location else None,
                            'longitude': resource.location.x if resource.location else None
                        }
                    },
                    'dates': {
                        'creation_date': resource.creation_date.isoformat() if resource.creation_date else None,
                        'last_update': resource.last_update.isoformat() if resource.last_update else None,
                        'created_at': resource.created_at.isoformat(),
                        'updated_at': resource.updated_at.isoformat()
                    },
                    'status': {
                        'is_active': resource.is_active
                    },
                    'contact': {
                        'phone': getattr(resource, 'contact_phone', None),
                        'email': getattr(resource, 'contact_email', None),
                        'website': getattr(resource, 'contact_website', None)
                    }
                }
                
                # Ajouter les données multilingues si disponibles
                if resource.multilingual_data:
                    resource_data['multilingual_data'] = resource.multilingual_data
                
                export_data['resources'].append(resource_data)
            
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_data, jsonfile, indent=2, ensure_ascii=False, default=str)
            
            file_size = os.path.getsize(file_path)
            logger.info(f"Export JSON créé: {file_path} ({file_size} bytes)")
            
            return file_path, file_size
            
        except Exception as e:
            logger.error(f"Erreur export JSON: {e}")
            raise
    
    def create_http_response(self, file_path: str, content_type: str) -> HttpResponse:
        """
        Crée une réponse HTTP pour télécharger un fichier
        
        Args:
            file_path: Chemin vers le fichier
            content_type: Type MIME du fichier
            
        Returns:
            HttpResponse pour le téléchargement
        """
        try:
            filename = os.path.basename(file_path)
            
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type=content_type)
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response['Content-Length'] = os.path.getsize(file_path)
                
            return response
            
        except Exception as e:
            logger.error(f"Erreur création réponse HTTP: {e}")
            raise
    
    def cleanup_old_exports(self, days_old: int = 7):
        """
        Nettoie les anciens fichiers d'export
        
        Args:
            days_old: Nombre de jours après lesquels supprimer les fichiers
        """
        try:
            cutoff_time = timezone.now() - timezone.timedelta(days=days_old)
            deleted_count = 0
            
            for filename in os.listdir(self.export_dir):
                file_path = os.path.join(self.export_dir, filename)
                
                if os.path.isfile(file_path):
                    file_time = datetime.fromtimestamp(os.path.getctime(file_path))
                    file_time = timezone.make_aware(file_time)
                    
                    if file_time < cutoff_time:
                        os.remove(file_path)
                        deleted_count += 1
                        logger.info(f"Fichier d'export supprimé: {filename}")
            
            logger.info(f"Nettoyage terminé: {deleted_count} fichiers supprimés")
            return deleted_count
            
        except Exception as e:
            logger.error(f"Erreur nettoyage exports: {e}")
            return 0
    
    def _get_localized_field(self, resource, field: str, language: str) -> str:
        """Récupère un champ localisé"""
        
        # Essayer d'abord la version localisée
        if resource.multilingual_data and language in resource.multilingual_data:
            localized_value = resource.multilingual_data[language].get(field)
            if localized_value:
                return localized_value
        
        # Fallback sur le champ principal
        return getattr(resource, field, '') or ''


# Classes pour les réponses d'export via API
class ExportResponse:
    """Classe utilitaire pour les réponses d'export"""
    
    @staticmethod
    def csv_response(queryset, filename=None, language='fr') -> HttpResponse:
        """Génère une réponse CSV directe"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'resources_export_{timestamp}.csv'
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # En-têtes
        headers = [
            'ID', 'Nom', 'Description', 'Types', 'Ville', 'Code Postal',
            'Latitude', 'Longitude', 'Date Création', 'Actif'
        ]
        writer.writerow(headers)
        
        # Données
        for resource in queryset:
            export_service = ExportService()
            row = [
                resource.resource_id,
                export_service._get_localized_field(resource, 'name', language),
                export_service._get_localized_field(resource, 'description', language),
                ', '.join(resource.resource_types or []),
                resource.city or '',
                resource.postal_code or '',
                resource.location.y if resource.location else '',
                resource.location.x if resource.location else '',
                resource.creation_date.strftime('%Y-%m-%d') if resource.creation_date else '',
                'Oui' if resource.is_active else 'Non'
            ]
            writer.writerow(row)
        
        return response